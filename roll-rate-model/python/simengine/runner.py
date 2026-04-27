from __future__ import annotations

import csv
import math
import os
import random
import sys
from collections import defaultdict
from typing import Any, Dict, List, Optional, Tuple

from .data_prep import (
    DataManager,
    calc,
    calc_dynamic,
    build_static_cache,
    classify_model_terms,
    get_dial,
    init_data_manager,
    init_time_varying_state,
    step_period_fields,
    load_loans,
    validate_loan,
    _get_registry,
)
from .dump import should_dump, new_collector, snap_pre, snap_post, write_csv

CF_COL = [
    "cnt", "begin_bal", "end_bal", "int_pmt", "sch_int", "prin_pmt", "sch_prin",
    "pif_bal", "net_recov", "loss", "liq_bal", "dq30", "dq60", "dq90", "dq120",
    "pif_cnt", "liq_cnt", "dq30_bal", "dq60_bal", "dq90_bal", "dq120_bal",
    "recov", "cost2srvc", "sfee_pmt", "bk", "cf", "cf_delta", "int_rate", "irr", "npv",
]
CF_DICT = {name: i for i, name in enumerate(CF_COL)}

META_COLS = ["loan_id", "term", "grade", "loan_age", "ofico", "fico_bucket", "term_fico", "platform_f", "orig_bal"]


def _stable_seed(seed0: int, loan_id: str, path: int) -> int:
    s = f"{seed0}|{loan_id}|{path}"
    h = 1469598103934665603
    for ch in s:
        h ^= ord(ch)
        h = (h * 1099511628211) & 0xFFFFFFFFFFFFFFFF
    return int(h % (2 ** 32))


def _build_transition_layout(dm: DataManager) -> Dict[str, Dict]:
    """Pre-compute per-status transition metadata (C++ TransitionLayout).

    For each from_status, stores:
      stay_idx, model_names (list), model_links (list of link|None),
      prob_col_indices (list mapping roll_to idx -> flat prob schema col)
    """
    layout = {}
    for from_s, roll_to in dm.status_to_roll.items():
        cf = dm.clean_status_dict.get(from_s, from_s)
        stay_idx = roll_to.index(from_s)
        model_names = []
        model_links = []
        for to_s in roll_to:
            if to_s == from_s:
                model_names.append(None)
                model_links.append(None)
            else:
                ct = dm.clean_status_dict.get(to_s, to_s)
                name = f"from{cf}_{ct}"
                model_names.append(name)
                model_links.append(dm.models.get(name))
        layout[from_s] = {
            "stay_idx": stay_idx,
            "model_names": model_names,
            "model_links": model_links,
            "prob_col_idx": dm.prob_layout.get(from_s, []),
        }
    return layout


def _softmax_transition(loan: Dict, from_status: str, roll_to: List[str],
                        dm: DataManager, dial_per: int,
                        rng: random.Random,
                        logit_cache: Optional[Dict[str, float]] = None,
                        tl: Optional[Dict] = None,
                        ) -> Tuple[str, List[float]]:
    """Returns (status_to, prob_final) where prob_final is indexed by roll_to."""
    if tl is None:
        tl = dm._transition_layout[from_status]
    stay_idx = tl["stay_idx"]
    model_names = tl["model_names"]
    model_links = tl["model_links"]

    scores = []
    for i, link in enumerate(model_links):
        if i == stay_idx:
            continue
        if link:
            mn = model_names[i]
            if logit_cache is not None and mn in logit_cache:
                z = logit_cache[mn] + calc_dynamic(link, loan)
            else:
                z = calc(link, loan)
            scores.append(math.exp(z))
        else:
            scores.append(0.0)

    tmp_all = 1.0 + sum(scores)

    n = len(roll_to)
    prob_proj = [0.0] * n
    j = 0
    nonstay_sum = 0.0
    for i in range(n):
        if i == stay_idx:
            continue
        p = scores[j] / tmp_all
        prob_proj[i] = p
        nonstay_sum += p
        j += 1

    prob_proj[stay_idx] = max(0.0, 1.0 - nonstay_sum)

    # Skip dial adjustment when no dials configured (C++ optimization)
    if dm.dial_data:
        prob_adj = [
            p * get_dial(dm.dial_data, from_status, roll_to[i], dial_per, loan=loan)
            for i, p in enumerate(prob_proj)
        ]
        s = sum(prob_adj)
        if s > 0:
            prob_final = [p / s for p in prob_adj]
        else:
            prob_final = [0.0] * n
            prob_final[stay_idx] = 1.0
    else:
        prob_final = prob_proj

    u = rng.random()
    cumulative = 0.0
    for i in range(n):
        cumulative += prob_final[i]
        if u <= cumulative:
            return roll_to[i], prob_final
    return roll_to[-1], prob_final


def _prob_final_to_dict(prob_final: List[float], from_status: str,
                       roll_to: List[str], dm: DataManager) -> Dict[str, float]:
    """Convert flat prob_final list to a string-keyed dict (for dump only)."""
    cf = dm.clean_status_dict.get(from_status, from_status)
    d = {}
    for i, to_s in enumerate(roll_to):
        if to_s == from_status:
            d[f"from{cf}_stay"] = prob_final[i]
        else:
            ct = dm.clean_status_dict.get(to_s, to_s)
            d[f"from{cf}_{ct}"] = prob_final[i]
    return d


def _compute_payments(num_pay: float, begin_bal: float,
                      pi_pmt: float, r_m: float):
    bal = begin_bal
    total_int = 0.0
    total_prin = 0.0
    for _ in range(int(num_pay)):
        int_pmt = bal * r_m
        prin_pmt = pi_pmt - int_pmt
        prin_pmt = max(0.0, min(prin_pmt, bal))
        bal -= prin_pmt
        total_int += int_pmt
        total_prin += prin_pmt
        if bal <= 0.01:
            bal = 0.0
            break
    return total_int, total_prin, bal


def run_cf_one(loan: Dict[str, Any], dm: DataManager, dup: int,
               seed: int, dump_collector=None) -> Tuple[List[List[float]], List[Tuple]]:
    """Returns (cf, prob_log) where prob_log entries are
    (begin_bal, from_status, prob_final_list) tuples."""
    rng = random.Random(seed)
    loan = dict(loan)
    ci = CF_DICT
    n_per = dm.n_per
    tl_map = dm._transition_layout

    cf = [[0.0] * len(CF_COL) for _ in range(n_per)]
    prob_log: List[Tuple] = []

    status = str(loan["status"])
    end_bal = float(loan["end_bal"])
    term = int(loan["term"])
    int_rate = float(loan["int_rate"])
    r_m = int_rate / 12.0 if int_rate < 1.0 else int_rate / 1200.0

    z = (1 + r_m) ** term
    pi_pmt = loan["end_bal"] * r_m * z / (z - 1) if abs(z - 1) > 1e-9 else loan["end_bal"] / term

    init_time_varying_state(loan)

    # Build static logit cache once per loan (C++ optimization)
    logit_cache = build_static_cache(dm.models, loan)

    # Pre-extract terminal statuses and macro state for inner loop
    terminal = dm.terminal_statuses
    has_macro = hasattr(dm, "_macro_state") and dm._macro_state
    ms = dm._macro_state if has_macro else None

    for per in range(n_per):
        loan_age = int(loan.get("loan_age", 0))
        if end_bal <= 0.1 or status in terminal or loan_age > 480:
            break

        begin_bal = end_bal
        cf[per][ci["begin_bal"]] = begin_bal

        # Apply per-variable macro overrides
        if has_macro:
            r_dt = loan.get("r_dt", "")
            ym = str(r_dt)[:7] if r_dt else ""
            # Calendar vars (cpi_inflator_36, cpi_inflator_12, etc.)
            if ms.get("calendar_table"):
                row = ms["calendar_table"].get(ym)
                if row:
                    ms["_last_calendar"] = row
                last = ms.get("_last_calendar")
                if last:
                    for var_name in ms["calendar_vars"]:
                        if var_name in last:
                            loan[var_name] = last[var_name]
            # Loan-specific vars (rate_incentive_ALL)
            if ms.get("fico_coupon") and "rate_incentive_ALL" in ms["active_vars"]:
                r_key = ym.replace("-", "")
                bkt = loan.get("_fico_bkt", "")
                if bkt:
                    coupon_r = ms["fico_coupon"].get(f"{r_key}|{bkt}")
                    if coupon_r is not None:
                        ms.setdefault("_last_coupon_r", {})[id(loan)] = coupon_r
                    coupon_r = ms.get("_last_coupon_r", {}).get(id(loan))
                    coupon_v = loan.get("_coupon_at_vintage")
                    if coupon_r is not None and coupon_v is not None:
                        loan["rate_incentive_ALL"] = round(coupon_r - coupon_v, 4)

        if dump_collector is not None:
            snap_pre(dump_collector, loan, per, status)

        roll_to = dm.status_to_roll.get(status, [status])
        tl = tl_map.get(status)
        status_to, prob_final = _softmax_transition(loan, status, roll_to, dm, per, rng,
                                                     logit_cache=logit_cache, tl=tl)
        prob_log.append((begin_bal, status, prob_final))

        num_pay = dm.pmt_matrix.get(status_to, {}).get(status, 0.0)
        int_paid, prin_paid, end_bal = _compute_payments(num_pay, begin_bal, pi_pmt, r_m)

        if status_to == "PIF":
            prin_paid = begin_bal
            int_paid = begin_bal * r_m
            end_bal = 0.0
            cf[per][ci["pif_cnt"]] = 1.0
            cf[per][ci["pif_bal"]] = begin_bal
        elif status_to == "LIQ":
            loss = begin_bal * dm.liq_severity
            recovery = begin_bal - loss
            end_bal = 0.0
            prin_paid = 0.0
            int_paid = 0.0
            cf[per][ci["liq_cnt"]] = 1.0
            cf[per][ci["liq_bal"]] = begin_bal
            cf[per][ci["loss"]] = loss
            cf[per][ci["net_recov"]] = recovery
            cf[per][ci["recov"]] = recovery
        else:
            dq = dm.dq_buckets.get(status_to)
            if dq:
                cf[per][ci[dq[0]]] = 1.0
                cf[per][ci[dq[1]]] = end_bal

        cf[per][ci["end_bal"]] = end_bal
        cf[per][ci["int_pmt"]] = int_paid
        cf[per][ci["prin_pmt"]] = prin_paid
        cf[per][ci["cnt"]] = 1.0

        # Scheduled interest & principal
        cf[per][ci["sch_int"]]  = begin_bal * r_m
        cf[per][ci["sch_prin"]] = max(0.0, pi_pmt - cf[per][ci["sch_int"]])

        if dump_collector is not None:
            # Build dict for dump only (not on hot path)
            prob_dict = _prob_final_to_dict(prob_final, status, roll_to, dm)
            snap_post(dump_collector, status_to, prob_dict,
                      begin_bal, end_bal, int_paid, prin_paid,
                      cf[per][ci["loss"]])

        # Advance age + prepare next period's context in one step
        step_period_fields(loan, per + 1)

        loan["end_bal"] = end_bal
        status = status_to

    if dup > 1:
        inv = 1.0 / dup
        for row in cf:
            for c in range(len(row)):
                row[c] *= inv

    return cf, prob_log


def _build_prob_schema(dm: DataManager) -> Tuple[List[str], Dict[str, List[int]]]:
    """Build a flat probability schema from the status-roll config.

    Returns (prob_keys, layout) where:
      prob_keys: sorted list of all possible probability column names
      layout: {from_status: [col_idx_for_each_roll_to_entry]}
    """
    all_keys = set()
    for from_s, roll_to in dm.status_to_roll.items():
        cf = dm.clean_status_dict.get(from_s, from_s)
        for to_s in roll_to:
            if to_s == from_s:
                all_keys.add(f"from{cf}_stay")
            else:
                ct = dm.clean_status_dict.get(to_s, to_s)
                all_keys.add(f"from{cf}_{ct}")
    prob_keys = sorted(all_keys)
    key_idx = {k: i for i, k in enumerate(prob_keys)}

    layout = {}
    for from_s, roll_to in dm.status_to_roll.items():
        cf = dm.clean_status_dict.get(from_s, from_s)
        col_indices = []
        for to_s in roll_to:
            if to_s == from_s:
                col_indices.append(key_idx[f"from{cf}_stay"])
            else:
                ct = dm.clean_status_dict.get(to_s, to_s)
                col_indices.append(key_idx[f"from{cf}_{ct}"])
        layout[from_s] = col_indices
    return prob_keys, layout


def _build_loan_meta(loan: Dict) -> Dict:
    loan_meta = {}
    for col in META_COLS:
        loan_meta[col] = loan.get(col, "")
    if not loan_meta.get("orig_bal"):
        loan_meta["orig_bal"] = loan.get("end_bal", 0)
    return loan_meta


def _run_one_loan(loan: Dict, dm: DataManager, n_per: int, dup: int,
                  seed0: int, dump_cfg=None) -> Dict:
    validate_loan(loan)
    loan_id = str(loan.get("loan_id", "unknown"))
    loan_meta = _build_loan_meta(loan)

    loan_cf_sum = [[0.0] * len(CF_COL) for _ in range(n_per)]
    loan_errors = []
    max_paths = dump_cfg["max_paths"] if dump_cfg else 0

    # Pre-aggregate prob data across paths using pre-built schema
    n_pk = len(dm.prob_keys)
    prob_weighted = [[0.0] * n_pk for _ in range(n_per)]
    prob_bal = [0.0] * n_per

    dump_entries = []
    for path in range(dup):
        seed = _stable_seed(seed0, loan_id, path)
        collector = new_collector() if path < max_paths else None
        try:
            cf, prob_log = run_cf_one(loan, dm, dup, seed,
                                      dump_collector=collector)
            for p in range(n_per):
                for c in range(len(CF_COL)):
                    loan_cf_sum[p][c] += cf[p][c]

            # Accumulate prob-weighted sums using flat tuples + pre-built layout
            for pi, (bb, from_s, pf) in enumerate(prob_log):
                prob_bal[pi] += bb
                pw_row = prob_weighted[pi]
                col_idx = dm.prob_layout.get(from_s)
                if col_idx:
                    for ri, ci_val in enumerate(col_idx):
                        pw_row[ci_val] += pf[ri] * bb

            if collector and collector["rows"]:
                dump_entries.append((loan_id, path, collector))
        except Exception as e:
            loan_errors.append(f"{loan_id}|{path}: {e}")

    return {"meta": loan_meta, "cf": loan_cf_sum,
            "prob_weighted": prob_weighted, "prob_bal": prob_bal,
            "errors": loan_errors, "dump_entries": dump_entries}


def run_simulation(loans: List[Dict], input_dir: str,
                   n_per: int = 360, dup: int = 1, seed0: int = 42,
                   liq_severity: float = 0.60, dial_name: str = "",
                   status_to_roll=None, config=None,
                   workers: int = 1, mode: str = "auto") -> Dict:

    dm = init_data_manager(
        input_dir, n_per=n_per, dial_name=dial_name,
        liq_severity=liq_severity, status_to_roll=status_to_roll,
        config=config,
    )

    # Build flat probability schema (avoids per-entry dict creation)
    prob_keys, prob_layout = _build_prob_schema(dm)
    dm.prob_keys = prob_keys
    dm.prob_layout = prob_layout
    dm.prob_key_idx = {k: i for i, k in enumerate(prob_keys)}

    # Classify model terms as static vs dynamic for logit cache
    reg = _get_registry()
    dynamic_vars = reg.time_varying_names() | reg.macro_names()
    classify_model_terms(dm.models, dynamic_vars)

    # Pre-compute transition layout (C++ TransitionLayout)
    dm._transition_layout = _build_transition_layout(dm)

    # Load per-variable macro data sources
    macro_cfg = (config or {}).get("macro", {})
    if macro_cfg:
        from .data_prep import load_cpi_lookup, load_fico_coupon_lookup
        macro_state = {"active_vars": set(), "calendar_vars": set()}
        for var_name, vcfg in macro_cfg.items():
            if not isinstance(vcfg, dict) or vcfg.get("mode") != "custom":
                continue
            path = vcfg.get("path", "")
            if not path or not os.path.isfile(path):
                continue
            macro_state["active_vars"].add(var_name)
            if var_name == "rate_incentive_ALL":
                if "fico_coupon" not in macro_state:
                    macro_state["fico_coupon"] = load_fico_coupon_lookup(path)
                    print(f"  Macro {var_name}: {len(macro_state['fico_coupon'])} entries from {path}")
            else:
                macro_state["calendar_vars"].add(var_name)
                if "calendar_table" not in macro_state:
                    from .data_prep import load_macro_csv
                    macro_state["calendar_table"] = load_macro_csv(path)
                    print(f"  Macro table: {len(macro_state['calendar_table'])} rows from {path}")
        if macro_state["active_vars"]:
            dm._macro_state = macro_state

    # ── Dump: run first N loans sequentially with debug output ────
    dump_cfg = should_dump(config)
    dump_results, dump_errors = [], []
    all_dump_entries = []
    remaining_loans = loans
    if dump_cfg:
        n = min(dump_cfg["max_loans"], len(loans))
        os.makedirs(dump_cfg["output_dir"], exist_ok=True)
        print(f"  Dump: {n} loans x {dump_cfg['max_paths']} paths -> {dump_cfg['output_dir']}")
        for loan in loans[:n]:
            result = _run_one_loan(loan, dm, n_per, dup, seed0, dump_cfg=dump_cfg)
            if not result["errors"]:
                dump_results.append(result)
            dump_errors.extend(result["errors"])
            all_dump_entries.extend(result.get("dump_entries", []))
        remaining_loans = loans[n:]
        if all_dump_entries:
            fpath = write_csv(dump_cfg["output_dir"], all_dump_entries)
            print(f"  dump: {fpath} ({len(all_dump_entries)} loan-paths)")

    # ── Main simulation ───────────────────────────────────────────
    if mode == "sequential":  loan_results, errors = _run_sequential(remaining_loans, dm, n_per, dup, seed0)
    elif mode == "pool":      loan_results, errors = _run_pool(remaining_loans, dm, n_per, dup, seed0, workers)
    elif mode == "ray":       loan_results, errors = _run_parallel(remaining_loans, dm, n_per, dup, seed0, workers)
    else:                     raise ValueError(f"Unknown mode: {mode!r}  (expected sequential|pool|ray)")

    loan_results = dump_results + loan_results
    errors = dump_errors + errors

    cf_sum = [[0.0] * len(CF_COL) for _ in range(n_per)]
    for lr in loan_results:
        for p in range(n_per):
            for c in range(len(CF_COL)):
                cf_sum[p][c] += lr["cf"][p][c]

    return {"cf_sum": cf_sum, "loan_results": loan_results, "errors": errors,
            "dm": dm, "cf_col": CF_COL, "n_per": n_per}


def _pool_worker(args):
    loan, dm, n_per, dup, seed0 = args
    return _run_one_loan(loan, dm, n_per, dup, seed0)


def _run_pool(loans, dm, n_per, dup, seed0, workers):
    from multiprocessing import Pool
    import time as _time

    t0 = _time.time()
    tasks = [(loan, dm, n_per, dup, seed0) for loan in loans]
    with Pool(processes=workers) as pool:
        results = pool.map(_pool_worker, tasks)
    t1 = _time.time()
    print(f"  Pool execution: {t1 - t0:.2f}s ({workers} workers)")

    loan_results = []
    errors = []
    for result in results:
        if not result["errors"]:
            loan_results.append(result)
        errors.extend(result["errors"])
    return loan_results, errors


def _run_sequential(loans, dm, n_per, dup, seed0):
    loan_results = []
    errors = []
    for loan in loans:
        result = _run_one_loan(loan, dm, n_per, dup, seed0)
        if not result["errors"]:
            loan_results.append(result)
        errors.extend(result["errors"])
    return loan_results, errors


def _run_parallel(loans, dm, n_per, dup, seed0, workers):
    import ray
    import time as _time

    t_ray_start = _time.time()
    pkg_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if not ray.is_initialized():
        ray.init(
            num_cpus=workers,
            log_to_driver=True,
            include_dashboard=False,
            # working_dir uploads the directory and sets it as cwd for workers
            # so `import simengine` works.  Avoid env_vars — it hangs on Windows.
            runtime_env={"working_dir": pkg_dir},
        )
    t_ray_init = _time.time()
    print(f"  Ray init: {t_ray_init - t_ray_start:.2f}s", flush=True)

    dm_ref = ray.put(dm)
    t_put = _time.time()
    print(f"  Ray put DataManager: {t_put - t_ray_init:.2f}s", flush=True)

    @ray.remote
    def _process_batch(batch, dm_ref, n_per, dup, seed0):
        results = []
        for loan in batch:
            results.append(_run_one_loan(loan, dm_ref, n_per, dup, seed0))
        return results

    batch_size = max(1, len(loans) // (workers * 4))
    batches = [loans[i:i + batch_size] for i in range(0, len(loans), batch_size)]
    print(f"  Dispatching {len(batches)} batches ({batch_size} loans/batch) to {workers} workers",
          flush=True)

    futures = [_process_batch.remote(batch, dm_ref, n_per, dup, seed0)
               for batch in batches]
    try:
        batch_results = ray.get(futures)
    finally:
        ray.shutdown()
    t_sim = _time.time()
    print(f"  Sim execution: {t_sim - t_put:.2f}s", flush=True)

    loan_results = []
    errors = []
    for batch in batch_results:
        for result in batch:
            if not result["errors"]:
                loan_results.append(result)
            errors.extend(result["errors"])

    return loan_results, errors


def compute_metrics(cf_rows: List[List[float]], orig_bal: float) -> List[Dict[str, float]]:
    ci = CF_DICT
    cum_loss = 0.0
    metrics = []
    for p, row in enumerate(cf_rows):
        bb = row[ci["begin_bal"]]
        sp = row[ci["sch_prin"]]
        pif_bal = row[ci["pif_bal"]]
        liq_bal = row[ci["liq_bal"]]
        loss = row[ci["loss"]]
        cum_loss += loss

        denom = bb - sp
        smm_prepay = min(pif_bal / denom, 1.0) if denom > 0.1 else 0.0
        cpr = 1.0 - (1.0 - smm_prepay) ** 12

        smm_default = liq_bal / bb if bb > 0.1 else 0.0
        cdr = 1.0 - (1.0 - smm_default) ** 12

        cgl = cum_loss / orig_bal if orig_bal > 0 else 0.0

        metrics.append({"period": p + 1, "cpr": cpr, "cdr": cdr, "cgl": cgl,
                        "begin_bal": bb, "pif_bal": pif_bal, "liq_bal": liq_bal,
                        "loss": loss, "cum_loss": cum_loss,
                        "dq30_bal": row[ci["dq30_bal"]], "dq60_bal": row[ci["dq60_bal"]],
                        "dq90_bal": row[ci["dq90_bal"]], "dq120_bal": row[ci["dq120_bal"]]})
    return metrics


def _collect_prob_keys(loan_results: List[Dict], dm: DataManager = None) -> List[str]:
    if dm is not None and hasattr(dm, "prob_keys"):
        return dm.prob_keys
    keys = set()
    for lr in loan_results:
        for k in lr.get("prob_keys", []):
            keys.add(k)
    return sorted(keys)


def aggregate_by_groups(
    loan_results: List[Dict],
    group_by: List[str],
    n_per: int,
    dm: DataManager = None,
) -> Dict:
    groups: Dict[str, List[Dict]] = defaultdict(list)

    for lr in loan_results:
        meta = lr["meta"]
        key_parts = []
        for g in group_by:
            val = meta.get(g, "")
            if val is None:
                val = ""
            key_parts.append(f"{g}={val}")
        key = "|".join(key_parts)
        groups[key].append(lr)

    prob_keys = _collect_prob_keys(loan_results, dm=dm)

    # Find max possible age across all loans
    max_orig_age = 0
    for lr in loan_results:
        la = int(lr["meta"].get("loan_age", 0) or 0)
        if la > max_orig_age:
            max_orig_age = la
    max_age = max_orig_age + n_per

    agg_results = {}
    for group_key, members in groups.items():
        # Age-indexed accumulators (not period-indexed)
        group_cf = [[0.0] * len(CF_COL) for _ in range(max_age)]
        group_orig_bal = 0.0
        prob_weighted = [[0.0] * len(prob_keys) for _ in range(max_age)]
        prob_bal_total = [0.0] * max_age

        for lr in members:
            ob = lr["meta"].get("orig_bal") or 0
            group_orig_bal += float(ob) if ob else 0
            orig_age = int(lr["meta"].get("loan_age", 0) or 0)

            # Accumulate cashflows by loan_age
            for p in range(n_per):
                age = orig_age + p
                if age >= max_age:
                    break
                for c in range(len(CF_COL)):
                    group_cf[age][c] += lr["cf"][p][c]

            # Accumulate pre-aggregated probs by loan_age
            lr_pw = lr.get("prob_weighted")
            lr_pb = lr.get("prob_bal")
            if lr_pw:
                n_pk = len(prob_keys)
                for pi in range(min(n_per, len(lr_pw))):
                    age = orig_age + pi
                    if age >= max_age:
                        break
                    prob_bal_total[age] += lr_pb[pi]
                    src = lr_pw[pi]
                    dst = prob_weighted[age]
                    for ki in range(n_pk):
                        dst[ki] += src[ki]

        metrics = compute_metrics(group_cf, group_orig_bal)
        for i, m in enumerate(metrics):
            age = i  # metrics index = age index
            if age < max_age and prob_bal_total[age] > 0:
                for ki, pk in enumerate(prob_keys):
                    m[pk] = prob_weighted[age][ki] / prob_bal_total[age]
            else:
                for pk in prob_keys:
                    m[pk] = 0.0

        agg_results[group_key] = {"metrics": metrics, "orig_bal": group_orig_bal}

    agg_results["_prob_keys"] = prob_keys
    return agg_results


def _parse_group_key(group_key: str) -> Dict[str, str]:
    parts = {}
    for token in group_key.split("|"):
        k, _, v = token.partition("=")
        parts[k] = v
    return parts


DEFAULT_GROUP_BY = ["term", "grade"]


def write_results_xlsx(result: Dict, output_path: str,
                       group_by: Optional[List[str]] = None):
    from openpyxl import Workbook

    if group_by is None:
        group_by = DEFAULT_GROUP_BY

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb = Workbook()

    cf_sum = result["cf_sum"]
    loan_results = result["loan_results"]
    n_per = result["n_per"]

    ws_port = wb.active
    ws_port.title = "Portfolio"
    ws_port.append(["period"] + CF_COL)
    for i, row in enumerate(cf_sum):
        ws_port.append([i + 1] + [round(v, 2) for v in row])

    total_orig_bal = sum(
        float(lr["meta"].get("orig_bal") or 0)
        for lr in loan_results
    )
    port_metrics = compute_metrics(cf_sum, total_orig_bal)
    ws_pm = wb.create_sheet("Metrics_Portfolio")
    ws_pm.append(["period", "cpr", "cdr", "cgl", "begin_bal", "pif_bal",
                   "liq_bal", "loss", "cum_loss"])
    for m in port_metrics:
        if m["begin_bal"] == 0 and m["cpr"] == 0 and m["cdr"] == 0:
            continue
        ws_pm.append([m["period"], round(m["cpr"], 6), round(m["cdr"], 6),
                       round(m["cgl"], 6), round(m["begin_bal"], 2),
                       round(m["pif_bal"], 2), round(m["liq_bal"], 2),
                       round(m["loss"], 2), round(m["cum_loss"], 2)])

    agg = aggregate_by_groups(loan_results, group_by, n_per, dm=result.get("dm"))
    prob_keys = agg.pop("_prob_keys", [])

    base_cols = ["loan_age", "cpr", "cdr", "cgl",
                 "begin_bal", "pif_bal", "liq_bal", "loss", "cum_loss",
                 "dq30_bal", "dq60_bal", "dq90_bal", "dq120_bal"]

    ws_agg = wb.create_sheet("Metrics_Grouped")
    ws_agg.append(group_by + base_cols + prob_keys)

    for group_key in sorted(k for k in agg if k != "_prob_keys"):
        group_data = agg[group_key]
        metrics = group_data["metrics"]
        parts = _parse_group_key(group_key)
        group_vals = [parts.get(g, "") for g in group_by]

        # Probs at age A predict the transition observed at age A+1.
        # Financials at age A reflect what was observed at age A.
        # metrics[i] corresponds to cf_array[i], period = i+1.
        # cf_array[i] = cashflows from transitions at age i,
        # whose outcome is observed at age i+1.
        # So: financials for loan_age N come from metrics[N-1].
        #     probs for loan_age N come from metrics[N].

        for age in range(len(metrics)):
            has_probs = metrics[age].get(prob_keys[0], 0.0) != 0.0 if prob_keys else False
            has_financials = (age >= 1
                and (age - 1) < len(metrics)
                and (metrics[age - 1]["begin_bal"] > 0.01
                     or metrics[age - 1]["cpr"] > 0
                     or metrics[age - 1]["cdr"] > 0))

            if not has_probs and not has_financials:
                continue

            prob_vals = [round(metrics[age].get(pk, 0.0), 6) for pk in prob_keys]

            if has_financials:
                m = metrics[age - 1]
                ws_agg.append(group_vals + [
                    age,
                    round(m["cpr"], 6), round(m["cdr"], 6),
                    round(m["cgl"], 6), round(m["begin_bal"], 2),
                    round(m["pif_bal"], 2), round(m["liq_bal"], 2),
                    round(m["loss"], 2), round(m["cum_loss"], 2),
                    round(m.get("dq30_bal", 0), 2), round(m.get("dq60_bal", 0), 2),
                    round(m.get("dq90_bal", 0), 2), round(m.get("dq120_bal", 0), 2),
                ] + prob_vals)
            else:
                # Probs-only row (e.g. age 0 for new loans)
                ws_agg.append(group_vals + [
                    age, None, None, None, None,
                    None, None, None, None,
                    None, None, None, None,
                ] + prob_vals)

    wb.save(output_path)


def write_cf_csv(cf_sum: List[List[float]], output_path: str):
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["period"] + CF_COL)
        for i, row in enumerate(cf_sum):
            writer.writerow([i + 1] + [f"{v:.6f}" for v in row])
