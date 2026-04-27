#!/usr/bin/env python3
"""Merge C++ simulation CSVs into a single XLSX file.

Usage: python tools/csvs_to_xlsx.py <output_dir> [output.xlsx]

Reads:
  <output_dir>/sim_results.csv       -> Portfolio sheet
  <output_dir>/metrics_portfolio.csv -> Metrics_Portfolio sheet
  <output_dir>/metrics_grouped.csv   -> Metrics_Grouped sheet

Writes:
  <output_dir>/sim_results.xlsx (or specified path)
"""

import csv
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font


def add_csv_sheet(wb, sheet_name, csv_path):
    """Read a CSV and add it as a sheet in the workbook."""
    if not os.path.exists(csv_path):
        print(f"  Skipping {sheet_name}: {csv_path} not found")
        return
    ws = wb.create_sheet(sheet_name)
    with open(csv_path) as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            # Try to convert numeric values
            converted = []
            for val in row:
                try:
                    if "." in val:
                        converted.append(float(val))
                    else:
                        converted.append(int(val))
                except ValueError:
                    converted.append(val)
            ws.append(converted)
            if i == 0:
                for c in ws[1]:
                    c.font = Font(bold=True)


def main():
    out_dir = sys.argv[1] if len(sys.argv) > 1 else "output/par_2026_1"
    xlsx_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(out_dir, "sim_results.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    # Find the main CF csv (could be sim_results.csv or cpp_dup100.csv etc.)
    cf_csv = None
    for name in ["sim_results.csv", "cpp_dup100.csv"]:
        p = os.path.join(out_dir, name)
        if os.path.exists(p):
            cf_csv = p
            break

    if cf_csv:
        add_csv_sheet(wb, "Portfolio", cf_csv)

    add_csv_sheet(wb, "Metrics_Portfolio", os.path.join(out_dir, "metrics_portfolio.csv"))
    add_csv_sheet(wb, "Metrics_Grouped", os.path.join(out_dir, "metrics_grouped.csv"))
    add_csv_sheet(wb, "Metrics_Grouped_Period", os.path.join(out_dir, "metrics_grouped_period.csv"))

    os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
    wb.save(xlsx_path)
    print(f"XLSX written: {xlsx_path} ({len(wb.sheetnames)} sheets: {wb.sheetnames})")

    # Clean up individual CSVs
    for name in ["sim_results.csv", "metrics_portfolio.csv", "metrics_grouped.csv", "metrics_grouped_period.csv"]:
        p = os.path.join(out_dir, name)
        if os.path.exists(p):
            os.remove(p)
            print(f"  Removed: {p}")


if __name__ == "__main__":
    main()
