"""format_excel.py - Template-driven Excel formatting tool.

Scan Excel files to generate formatting templates, then apply them for
consistent, management-quality output.

Usage:
    # Scan file structure
    python format_excel.py input.xlsx --scan --header-row 3

    # Apply saved template
    python format_excel.py input.xlsx -t template.json -o output.xlsx

    # Apply in-place (creates .bak backup)
    python format_excel.py input.xlsx -t template.json --inplace

All column keys in templates reference ORIGINAL header names in the input file.
Renaming is applied as the last visual step.
"""

from __future__ import annotations

import argparse
import datetime
import json
import shutil
import sys
import tempfile
import warnings
from pathlib import Path
from statistics import median
from typing import Any

# Suppress openpyxl warnings about unsupported Excel extensions (e.g.,
# Microsoft-specific conditional formatting).  These are harmless -- openpyxl
# simply drops the proprietary extensions on read; our formatting still works.
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_DATE_FMT_HINTS = ("yy", "mm/d", "d/m", "yyyy")
_PCT_HEADER_KEYWORDS = ("rate", "pct", "%", "percent")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _build_col_map(ws, header_row: int) -> dict[str, list[int]]:
    """Map header names -> list of 1-based column indices.

    Handles duplicate headers by collecting all matching indices.
    Blank headers are silently skipped.
    """
    col_map: dict[str, list[int]] = {}
    for cell in ws[header_row]:
        name = cell.value
        if name is None or (isinstance(name, str) and not name.strip()):
            continue
        name = str(name).strip()
        col_map.setdefault(name, []).append(cell.column)
    return col_map


def _data_col_bounds(ws, header_row: int) -> tuple[int, int]:
    """Return (first_col, last_col) of columns that have a header value.

    Columns without a header (e.g., blank column A) are excluded from
    all styling operations.
    """
    first_col = ws.max_column
    last_col = 1
    for cell in ws[header_row]:
        if cell.value is not None:
            first_col = min(first_col, cell.column)
            last_col = max(last_col, cell.column)
    # Also check super-header row
    if header_row > 1:
        for cell in ws[header_row - 1]:
            if cell.value is not None:
                first_col = min(first_col, cell.column)
                last_col = max(last_col, cell.column)
    return first_col, last_col


def _estimate_formatted_width(value: Any, number_format: str) -> int:
    """Best-effort estimate of display character width for a cell."""
    if value is None:
        return 0
    if isinstance(value, bool):
        return 5
    if isinstance(value, datetime.datetime):
        return 11  # "mm/dd/yyyy"
    if isinstance(value, (int, float)):
        nf = number_format or "General"
        try:
            if nf == "General":
                return len(f"{value:.2f}")
            # Extract decimal places from patterns like "0.00", "#,##0.0"
            dp = 0
            if "." in nf:
                after_dot = nf.split(".")[-1]
                dp = sum(1 for c in after_dot if c in "0#")
                dp = min(dp, 6)
            has_comma = "#,#" in nf or ",##" in nf
            if has_comma:
                return len(f"{value:,.{dp}f}")
            return len(f"{value:.{dp}f}")
        except (ValueError, TypeError):
            pass
    return min(len(str(value)), 30)


def _hide_zero_columns(
    ws, ws_data, header_row: int,
    col_bounds: tuple[int, int] | None = None,
) -> None:
    """Hide columns where every data value is exactly 0 (or empty).

    Uses ws_data for computed values. Hides the column in ws (preserves data).
    """
    first_col, last_col = col_bounds or (1, ws.max_column)
    for c in range(first_col, last_col + 1):
        all_zero = True
        for r in range(header_row + 1, ws_data.max_row + 1):
            val = ws_data.cell(row=r, column=c).value
            if val is None:
                continue
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                if val != 0:
                    all_zero = False
                    break
            else:
                # Non-numeric value means this isn't a zero-only column
                all_zero = False
                break
        if all_zero:
            letter = get_column_letter(c)
            ws.column_dimensions[letter].hidden = True


# ---------------------------------------------------------------------------
# Scan
# ---------------------------------------------------------------------------


def _detect_column_type(
    ws, col_idx: int, header_row: int, header_name: str | None,
) -> dict[str, Any]:
    """Analyze a column's data and return type / stats / suggested format."""
    nums: list[float] = []
    texts: list[str] = []
    date_count = 0
    pct_format_count = 0
    total = 0

    for row_idx in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        val = cell.value
        if val is None:
            continue
        total += 1
        nf = (cell.number_format or "").lower()

        # Date detection via type or format string
        if isinstance(val, datetime.datetime):
            date_count += 1
            continue
        if isinstance(val, (int, float)) and any(h in nf for h in _DATE_FMT_HINTS):
            date_count += 1
            continue

        # Percentage format on the cell itself
        if "%" in nf:
            pct_format_count += 1

        if isinstance(val, (int, float)) and not isinstance(val, bool):
            nums.append(float(val))
        elif isinstance(val, str):
            texts.append(val)

    if total == 0:
        return {"detected_type": "empty", "count": 0}

    # Majority dates
    if date_count > total * 0.5:
        return {
            "detected_type": "date",
            "count": date_count,
            "suggested_format": "mm/dd/yyyy",
        }

    # Mostly text
    if not nums:
        return {
            "detected_type": "text",
            "count": len(texts),
            "sample_values": texts[:5],
        }

    # Numeric analysis
    abs_nonzero = [abs(v) for v in nums if v != 0]
    med_abs = median(abs_nonzero) if abs_nonzero else 0.0
    all_int = all(v == int(v) for v in nums)

    # Conservative percentage check
    header_lower = (header_name or "").lower()
    has_pct_kw = any(kw in header_lower for kw in _PCT_HEADER_KEYWORDS)
    all_small = abs_nonzero and all(v <= 1.0 for v in abs_nonzero)

    if pct_format_count > len(nums) * 0.5:
        dtype, fmt = "percentage", "0.00%"
    elif all_small and has_pct_kw:
        dtype, fmt = "likely_percentage", "0.00%"
    elif all_int:
        dtype, fmt = "integer", "#,##0"
    elif med_abs >= 100:
        dtype, fmt = "float", "#,##0"
    elif med_abs >= 10:
        dtype, fmt = "float", "#,##0.0"
    else:
        dtype, fmt = "float", "0.00"

    result: dict[str, Any] = {
        "detected_type": dtype,
        "count": len(nums),
        "stats": {
            "min": round(min(nums), 4),
            "max": round(max(nums), 4),
            "median_abs": round(med_abs, 4),
        },
        "suggested_format": fmt,
    }
    if all_int:
        result["all_integer"] = True
    return result


def scan_workbook(path: str, header_row: int = 1) -> dict[str, Any]:
    """Scan an Excel file and return a report with a draft template.

    Returns:
        dict with keys: file, header_row, sheets (detailed), draft_template.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    sheets_report: list[dict] = []
    draft_col_fmts: dict[str, str] = {}

    for ws in wb:
        cols_info: list[dict] = []

        for col_idx in range(1, ws.max_column + 1):
            hdr_cell = ws.cell(row=header_row, column=col_idx)
            hdr = str(hdr_cell.value).strip() if hdr_cell.value else None

            info = _detect_column_type(ws, col_idx, header_row, hdr)
            info["index"] = col_idx
            info["header"] = hdr
            cols_info.append(info)

            # Populate draft formats (skip likely_percentage - needs confirmation)
            if (
                hdr
                and "suggested_format" in info
                and info["detected_type"] != "likely_percentage"
                and hdr not in draft_col_fmts
            ):
                draft_col_fmts[hdr] = info["suggested_format"]

        sheets_report.append({
            "name": ws.title,
            "rows": ws.max_row,
            "cols": ws.max_column,
            "columns": cols_info,
        })

    wb.close()

    draft_template: dict[str, Any] = {
        "name": "auto_generated",
        "sheets": ["*"],
        "exclude_sheets": [],
        "header_row": header_row,
        "header_style": {
            "font_bold": True,
            "font_color": "#FFFFFF",
            "fill_color": "#305496",
            "alignment": "center",
            "freeze": True,
            "auto_filter": True,
        },
        "column_formats": draft_col_fmts,
        "magnitude_format": {
            "enabled": True,
            "priority": "fallback",
            "rules": [
                {"min_abs": 100, "format": "#,##0"},
                {"min_abs": 10, "format": "#,##0.0"},
                {"min_abs": 0, "format": "0.00"},
            ],
        },
        "col_width": "auto",
    }

    return {
        "file": str(path),
        "header_row": header_row,
        "sheets": sheets_report,
        "draft_template": draft_template,
    }


# ---------------------------------------------------------------------------
# Apply helpers
# ---------------------------------------------------------------------------


def _resolve_sheets(wb, template: dict) -> list[str]:
    """Return sheet names to process: sheets selection minus exclude_sheets."""
    sheets_spec = template.get("sheets", ["*"])
    exclude = set(template.get("exclude_sheets", []))

    if sheets_spec == ["*"]:
        targets = list(wb.sheetnames)
    else:
        targets = [s for s in sheets_spec if s in wb.sheetnames]

    return [s for s in targets if s not in exclude]


def apply_number_formats(
    ws, ws_data, col_map: dict[str, list[int]],
    header_row: int, template: dict,
) -> None:
    """Apply column_formats (explicit) then magnitude_format (fallback).

    Uses ws for editing, ws_data for computed values (magnitude sampling).
    """
    col_formats = template.get("column_formats", {})
    mag_cfg = template.get("magnitude_format", {})
    mag_enabled = mag_cfg.get("enabled", False)
    mag_rules = mag_cfg.get("rules", [])

    formatted_cols: set[int] = set()

    # --- Explicit column_formats (highest priority) ---
    for col_name, fmt in col_formats.items():
        for col_idx in col_map.get(col_name, []):
            formatted_cols.add(col_idx)
            for r in range(header_row + 1, ws.max_row + 1):
                ws.cell(row=r, column=col_idx).number_format = fmt

    # --- Magnitude fallback for remaining numeric columns ---
    if not mag_enabled:
        return

    for _col_name, indices in col_map.items():
        for col_idx in indices:
            if col_idx in formatted_cols:
                continue

            # Sample computed values from ws_data
            abs_vals: list[float] = []
            has_date = False
            for r in range(header_row + 1, ws_data.max_row + 1):
                val = ws_data.cell(row=r, column=col_idx).value
                if isinstance(val, datetime.datetime):
                    has_date = True
                    break
                if (
                    isinstance(val, (int, float))
                    and not isinstance(val, bool)
                    and val != 0
                ):
                    abs_vals.append(abs(val))

            if has_date or not abs_vals:
                continue

            med = median(abs_vals)
            chosen_fmt = None
            for rule in mag_rules:
                if med >= rule["min_abs"]:
                    chosen_fmt = rule["format"]
                    break

            if chosen_fmt:
                for r in range(header_row + 1, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx).number_format = chosen_fmt


def _build_style_objects(cfg: dict) -> tuple:
    """Build Font, PatternFill, Alignment, Border from a style config dict.

    Shared by header_style and super_header_style.
    Returns (font, fill, alignment, border) -- any may be None.
    """
    font_kw: dict[str, Any] = {}
    if cfg.get("font_bold"):
        font_kw["bold"] = True
    if cfg.get("font_color"):
        font_kw["color"] = cfg["font_color"].lstrip("#")
    if cfg.get("font_size"):
        font_kw["size"] = cfg["font_size"]
    font = Font(**font_kw) if font_kw else None

    fill = None
    if cfg.get("fill_color"):
        c = cfg["fill_color"].lstrip("#")
        fill = PatternFill(start_color=c, end_color=c, fill_type="solid")

    alignment = None
    if cfg.get("alignment"):
        alignment = Alignment(
            horizontal=cfg["alignment"], vertical="center", wrap_text=False,
        )

    border_style = cfg.get("border_style", "thin")
    border_color = cfg.get("border_color", "000000").lstrip("#")
    bottom_side = Side(style=border_style, color=border_color)
    thin_side = Side(style="thin", color=border_color)
    border = Border(
        bottom=bottom_side, top=thin_side, left=thin_side, right=thin_side,
    )

    return font, fill, alignment, border


def apply_header_style(ws, header_row: int, template: dict) -> None:
    """Style header row and optional super-header rows above it."""
    # --- Super-header rows (rows above header_row) ---
    super_cfg = template.get("super_header_style")
    if super_cfg and header_row > 1:
        s_font, s_fill, s_align, s_border = _build_style_objects(super_cfg)
        for row_idx in range(1, header_row):
            for cell in ws[row_idx]:
                if cell.value is None:
                    continue
                if s_font:
                    cell.font = s_font
                if s_fill:
                    cell.fill = s_fill
                if s_align:
                    cell.alignment = s_align
                cell.border = s_border

    # --- Main header row ---
    cfg = template.get("header_style")
    if not cfg:
        return

    font, fill, alignment, header_border = _build_style_objects(cfg)

    for cell in ws[header_row]:
        if cell.value is None:
            continue
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        cell.border = header_border

    if cfg.get("freeze"):
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        # Reset sheetView scroll position so the file opens at the top
        # (source files may have been saved scrolled to an arbitrary row)
        if ws.views.sheetView:
            ws.views.sheetView[0].topLeftCell = "A1"

    if cfg.get("auto_filter"):
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"

    # --- Per-section colors (override fill for super-header + header rows) ---
    sec_colors = template.get("section_colors")
    dividers = template.get("section_dividers", [])
    if sec_colors and dividers:
        # Build column -> color index mapping using data bounds
        # Find first column with a header to avoid coloring blank cols (A, B)
        first_data_col = ws.max_column
        for cell in ws[header_row]:
            if cell.value is not None:
                first_data_col = min(first_data_col, cell.column)
                break
        boundaries = [first_data_col] + dividers + [ws.max_column + 1]
        for sec_idx in range(len(boundaries) - 1):
            if sec_idx >= len(sec_colors):
                break
            c_hex = sec_colors[sec_idx].lstrip("#")
            sec_fill = PatternFill(
                start_color=c_hex, end_color=c_hex, fill_type="solid",
            )
            col_start = boundaries[sec_idx]
            col_end = boundaries[sec_idx + 1]
            # Apply to super-header rows and header row
            for row_idx in range(max(1, header_row - 1), header_row + 1):
                for col_idx in range(col_start, col_end):
                    if col_idx > ws.max_column:
                        break
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = sec_fill


def _build_row_fills(
    ws_data, header_row: int, template: dict,
) -> dict[int, PatternFill | None]:
    """Build a mapping of row -> fill color for data rows.

    Supports two modes:
      - group_by_column: alternating colors per group (e.g., Asset Subtype).
        Automatically cycles through a palette. Scalable to any number of groups.
      - banded_rows: simple alternating single color.
    If neither is set, returns empty dict (no row fills).
    """
    row_fills: dict[int, PatternFill | None] = {}
    group_cfg = template.get("group_by_column")

    if group_cfg:
        col_name = group_cfg.get("column")
        palette = group_cfg.get("colors", [
            "#E8EDF2", "#E7F0E5", "#FDF2E9", "#F0E6EF",
            "#E5ECF0", "#F5EADF", "#E6EDE8", "#F2E8E8",
        ])
        # Find the column index by header name
        group_col = None
        for cell in ws_data.parent[ws_data.title][header_row]:
            if cell.value and str(cell.value).strip() == col_name:
                group_col = cell.column
                break
        if not group_col:
            # Fall back: try ws_data directly
            for c in range(1, ws_data.max_column + 1):
                if ws_data.cell(header_row, c).value and \
                   str(ws_data.cell(header_row, c).value).strip() == col_name:
                    group_col = c
                    break

        if group_col:
            # Detect groups in order and assign colors
            seen: dict[str, int] = {}
            group_idx = -1
            for r in range(header_row + 1, ws_data.max_row + 1):
                val = ws_data.cell(row=r, column=group_col).value
                key = str(val).strip() if val else ""
                if key not in seen:
                    group_idx += 1
                    seen[key] = group_idx
                cidx = seen[key] % len(palette)
                c_hex = palette[cidx].lstrip("#")
                row_fills[r] = PatternFill(
                    start_color=c_hex, end_color=c_hex, fill_type="solid",
                )
            return row_fills

    # Fallback: simple banded rows
    banded_cfg = template.get("banded_rows")
    if banded_cfg:
        bc = banded_cfg.get("color", "#F2F2F2").lstrip("#")
        band_fill = PatternFill(start_color=bc, end_color=bc, fill_type="solid")
        for r in range(header_row + 1, ws_data.max_row + 1):
            if (r - header_row) % 2 == 0:
                row_fills[r] = band_fill

    return row_fills


def apply_data_style(
    ws, ws_data, header_row: int, template: dict,
    col_bounds: tuple[int, int] | None = None,
) -> None:
    """Apply borders, row coloring (group or banded), and number alignment.

    Only applies to columns within col_bounds (first_col, last_col).
    """
    first_col, last_col = col_bounds or (1, ws.max_column)
    border_cfg = template.get("borders")

    # Pre-build border
    cell_border = None
    if border_cfg:
        color = border_cfg.get("color", "#D9D9D9").lstrip("#")
        style = border_cfg.get("style", "thin")
        side = Side(style=style, color=color)
        cell_border = Border(top=side, bottom=side, left=side, right=side)

    # Row fills (group-based or banded)
    row_fills = _build_row_fills(ws_data, header_row, template)

    # Number alignment (right-align numbers, left-align text)
    num_align = Alignment(horizontal="right")
    text_align = Alignment(horizontal="left")

    for r in range(header_row + 1, ws.max_row + 1):
        row_fill = row_fills.get(r)
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            data_val = ws_data.cell(row=r, column=c).value

            # Borders on all cells in the data range
            if cell_border:
                cell.border = cell_border

            # Row fill (group color or banded)
            if row_fill:
                cell.fill = row_fill

            # Alignment based on data type
            if data_val is not None:
                if isinstance(data_val, (int, float)) and not isinstance(data_val, bool):
                    cell.alignment = num_align
                elif isinstance(data_val, str):
                    cell.alignment = text_align


def apply_section_dividers(ws, header_row: int, template: dict) -> None:
    """Add thick left borders at section boundaries to separate New/Prod/Diff.

    section_dividers: list of 1-based column indices where a section starts.
    A medium left border is drawn from the super-header row down to the last data row.
    """
    dividers = template.get("section_dividers")
    if not dividers:
        return

    div_color = template.get("section_divider_color", "#4472C4").lstrip("#")
    div_side = Side(style="medium", color=div_color)

    # Apply from row 1 (or row before header) through all data rows
    start_row = max(1, header_row - 1)  # include super-header row
    for col_idx in dividers:
        if col_idx < 1 or col_idx > ws.max_column:
            continue
        for r in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            # Preserve existing border sides, upgrade left to divider
            old = cell.border
            cell.border = Border(
                left=div_side,
                right=old.right,
                top=old.top,
                bottom=old.bottom,
            )


def apply_outer_border(
    ws, header_row: int, template: dict,
    col_bounds: tuple[int, int] | None = None,
) -> None:
    """Draw a medium border around the entire table (super-headers through last row)."""
    cfg = template.get("outer_border")
    if not cfg:
        return

    first_col, last_col = col_bounds or (1, ws.max_column)
    color = cfg.get("color", "#000000").lstrip("#")
    style = cfg.get("style", "medium")
    outer = Side(style=style, color=color)

    top_row = max(1, header_row - 1)  # include super-header row
    bot_row = ws.max_row

    for r in range(top_row, bot_row + 1):
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            old = cell.border

            new_top = outer if r == top_row else old.top
            new_bot = outer if r == bot_row else old.bottom
            new_left = outer if c == first_col else old.left
            new_right = outer if c == last_col else old.right

            # Only update if this cell is on the perimeter
            if r == top_row or r == bot_row or c == first_col or c == last_col:
                cell.border = Border(
                    top=new_top, bottom=new_bot,
                    left=new_left, right=new_right,
                )


def apply_conditional_formatting(
    ws, col_map: dict[str, list[int]], header_row: int, template: dict,
) -> None:
    """Apply conditional formatting rules (color scales, etc.).

    Target columns via col_indices (1-based, takes precedence) or columns (name list).
    """
    rules = template.get("conditional_format", [])
    data_start = header_row + 1
    data_end = ws.max_row

    for rule in rules:
        # Resolve target columns
        if rule.get("col_indices"):
            targets = [i for i in rule["col_indices"] if 1 <= i <= ws.max_column]
        elif rule.get("columns"):
            targets = []
            for name in rule["columns"]:
                targets.extend(col_map.get(name, []))
        else:
            continue

        rtype = rule.get("type", "3_color_scale")

        if rtype == "3_color_scale":
            for col_idx in targets:
                letter = get_column_letter(col_idx)
                rng = f"{letter}{data_start}:{letter}{data_end}"
                cs = ColorScaleRule(
                    start_type="min",
                    start_color=rule.get("min_color", "F8696B").lstrip("#"),
                    mid_type="percentile",
                    mid_value=50,
                    mid_color=rule.get("mid_color", "FFFFFF").lstrip("#"),
                    end_type="max",
                    end_color=rule.get("max_color", "63BE7B").lstrip("#"),
                )
                ws.conditional_formatting.add(rng, cs)


def apply_column_renames(
    ws, header_row: int, col_map: dict[str, list[int]], template: dict,
) -> None:
    """Rename column headers. Uses original col_map; runs before auto-widths."""
    renames = template.get("column_rename", {})
    for old_name, new_name in renames.items():
        for col_idx in col_map.get(old_name, []):
            ws.cell(row=header_row, column=col_idx).value = new_name


def auto_column_widths(
    ws, ws_data, header_row: int,
    col_bounds: tuple[int, int] | None = None,
) -> None:
    """Set column widths from content. Runs AFTER renames.

    Uses ws for header text (post-rename) and ws_data for computed data values,
    combined with the number formats already applied to ws.
    """
    first_col, last_col = col_bounds or (1, ws.max_column)
    for col_idx in range(first_col, last_col + 1):
        max_w = 0
        letter = get_column_letter(col_idx)

        # Header width (after renames)
        hdr = ws.cell(row=header_row, column=col_idx).value
        if hdr:
            max_w = len(str(hdr))

        # Data width (sample up to 100 rows for performance)
        end_row = min(ws.max_row, header_row + 100)
        for r in range(header_row + 1, end_row + 1):
            val = ws_data.cell(row=r, column=col_idx).value
            if val is None:
                continue
            nf = ws.cell(row=r, column=col_idx).number_format
            max_w = max(max_w, _estimate_formatted_width(val, nf))

        width = max(min(max_w + 2, 30), 8)
        ws.column_dimensions[letter].width = width


# ---------------------------------------------------------------------------
# Main apply orchestrator
# ---------------------------------------------------------------------------


def apply_template(
    path: str,
    template: dict,
    output_path: str | None = None,
    inplace: bool = False,
) -> str:
    """Apply a formatting template to an Excel file.

    Returns the path of the output file.
    """
    src = Path(path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    # Backup for --inplace (before any modification)
    if inplace:
        bak = src.parent / (src.name + ".bak")
        shutil.copy2(src, bak)
        print(f"Backup: {bak}", file=sys.stderr)

    # Load twice: data_only for value sampling, normal for editing
    wb_data = openpyxl.load_workbook(src, data_only=True)
    wb = openpyxl.load_workbook(src)

    header_row = template.get("header_row", 1)
    sheets = _resolve_sheets(wb, template)

    if not sheets:
        print("Warning: no sheets matched template spec.", file=sys.stderr)

    for name in sheets:
        ws = wb[name]
        ws_data = wb_data[name]
        col_map = _build_col_map(ws, header_row)
        col_bounds = _data_col_bounds(ws, header_row)

        # Order of operations:
        # Clear literal "NULL" strings (data cleanup before any formatting)
        first_col, last_col = col_bounds
        for r in range(header_row + 1, ws.max_row + 1):
            for c in range(first_col, last_col + 1):
                if ws.cell(row=r, column=c).value == "NULL":
                    ws.cell(row=r, column=c).value = None

        # 0. Hide all-zero columns (before any formatting)
        if template.get("hide_zero_columns"):
            _hide_zero_columns(ws, ws_data, header_row, col_bounds)
        # 1. Number formats (explicit, then magnitude fallback)
        apply_number_formats(ws, ws_data, col_map, header_row, template)
        # 2. Header + super-header style
        apply_header_style(ws, header_row, template)
        # 3. Data style (borders, group/banded rows, number alignment)
        apply_data_style(ws, ws_data, header_row, template, col_bounds)
        # 4. Section dividers (thick left borders between New/Prod/Diff)
        apply_section_dividers(ws, header_row, template)
        # 5. Outer border around entire table
        apply_outer_border(ws, header_row, template, col_bounds)
        # 6. Conditional formatting
        apply_conditional_formatting(ws, col_map, header_row, template)
        # 7. Column renames (visual only, uses original col_map)
        apply_column_renames(ws, header_row, col_map, template)
        # 8. Auto column widths (AFTER renames so widths fit new names)
        if template.get("col_width") == "auto":
            auto_column_widths(ws, ws_data, header_row, col_bounds)

    wb_data.close()

    # Remove non-target sheets so output contains only formatted sheets
    for name in list(wb.sheetnames):
        if name not in sheets:
            del wb[name]

    # Save output
    if inplace:
        # Atomic write: temp file -> replace original
        tmp = tempfile.NamedTemporaryFile(
            dir=src.parent, suffix=".xlsx", delete=False,
        )
        tmp_path = Path(tmp.name)
        tmp.close()
        try:
            wb.save(tmp_path)
            wb.close()
            tmp_path.replace(src)
            return str(src)
        except Exception:
            tmp_path.unlink(missing_ok=True)
            raise
    else:
        out = Path(output_path) if output_path else (
            src.parent / f"{src.stem}_formatted{src.suffix}"
        )
        wb.save(out)
        wb.close()
        return str(out)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> None:
    p = argparse.ArgumentParser(
        description="Template-driven Excel formatting tool.",
        epilog="Column keys in templates always refer to original header names.",
    )
    p.add_argument("input", help="Path to the Excel file")
    p.add_argument(
        "--scan", action="store_true",
        help="Scan file and print report + draft template JSON to stdout",
    )
    p.add_argument(
        "--header-row", type=int, default=1,
        help="1-based header row number (default: 1)",
    )
    p.add_argument("-t", "--template", help="JSON template file to apply")
    p.add_argument("-o", "--output", help="Output file path")
    p.add_argument(
        "--inplace", action="store_true",
        help="Overwrite input file (creates .bak backup first)",
    )

    args = p.parse_args()

    if args.scan:
        report = scan_workbook(args.input, header_row=args.header_row)
        json.dump(report, sys.stdout, indent=2, default=str)
        sys.stdout.write("\n")
        return

    if not args.template:
        p.error("Provide --scan or -t/--template.")

    with open(args.template) as f:
        tmpl = json.load(f)

    # CLI --header-row overrides template if explicitly provided
    if args.header_row != 1:
        tmpl["header_row"] = args.header_row

    if not args.output and not args.inplace:
        p.error("Provide -o/--output or --inplace.")

    out = apply_template(
        args.input, tmpl, output_path=args.output, inplace=args.inplace,
    )
    print(f"Done: {out}", file=sys.stderr)


if __name__ == "__main__":
    main()


# python formatter/format_excel.py formatter/risk_diff_crt.xlsx -t formatter/templates/risk_diff_base.json -o formatter/risk_diff_crt_formatted.xlsx