from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from dial_utils import (
    extract_summary_rows,
    find_latest_tracking_file,
    normalize_error_window,
)


# ==============================
# CONFIGURATION
# ==============================
BASE_DIRS = {
    "Dialed": Path(r"R:\QR\Resi_shared\tracking\Dialed"),
    "Undialed": Path(r"R:\QR\Resi_shared\tracking\Undialed"),
}

DEALTYPES = [
    "STACR",
    "CAS",
    "JUMBO",
    "HE",
    "NONQM",

]

# Bucket types: 'WAC', 'AGE', 'FICO', or any other bucket type in your data
BUCKET_TYPE = "WAC"

# Error window selector: '3M', '6M', '12M' (set to None to show all error columns)
ERROR_WINDOW = "6M"

# Sheets with NO bucket sections: collect every Status row instead
STATUS_SHEETS = {"M30", "M60", "M90P", "M270P", "FCLS", "REO"}

# Sheets to exclude entirely
EXCLUDE_SHEETS = {"CDR", "CPR"}

# Optional manual overrides by report label (set to None to use latest)
MANUAL_EXCEL_PATHS = {
    "Dialed": None,
    "Undialed": None,
}

OUTPUT_DIR = Path("dial/outputs")
# ==============================


def _apply_excel_formatting(
    ws,
    df: pd.DataFrame,
    diff_col: str,
    sheet_title: str | None = None,
) -> None:
    data_start_row = 2

    # Optional sheet title row
    if sheet_title:
        ws.insert_rows(1)
        data_start_row = 3
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = sheet_title
        title_cell.font = Font(bold=True, color="FFFFFF", size=13)
        title_cell.fill = PatternFill("solid", fgColor="1F4E78")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.border = Border(
            left=Side(border_style="thin", color="D9D9D9"),
            right=Side(border_style="thin", color="D9D9D9"),
            top=Side(border_style="thin", color="D9D9D9"),
            bottom=Side(border_style="thin", color="D9D9D9"),
        )
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="D9D9D9")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[data_start_row - 1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border

    # Freeze top row and add filter
    ws.freeze_panes = f"A{data_start_row}"
    ws.auto_filter.ref = f"A{data_start_row - 1}:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Column widths
    for col_idx, col_name in enumerate(df.columns, start=1):
        series = df[col_name].astype(str).head(50)
        max_len = max(len(str(col_name)), *(len(v) for v in series))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(22, max(10, max_len + 2))

    # Number formats + alignment + borders + banded rows
    num_fmt = {
        "Avg Bal": "#,##0",
        "Loan Num": "#,##0",
        "MV(MM)": "#,##0.00",
        "WALA": "0.00",
        "WAC": "0.00",
        "FICO": "0",
        "OCLTV": "0.00",
        "Model_Dialed": "0.00",
        "Model_Undialed": "0.00",
        "Dialed Model Ratio": "0.000",
        "Current_Dial": "0.000",
        "Actual": "0.000",
        "Implied_Dial": "0.000",
        "Proposed_Dial": "0.000",
        diff_col: "0.00",
    }
    band_fill = PatternFill("solid", fgColor="F7F7F7")
    for row_idx in range(data_start_row, ws.max_row + 1):
        is_banded = (row_idx % 2 == 0)
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name in num_fmt:
                cell.number_format = num_fmt[col_name]
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if is_banded:
                cell.fill = band_fill
            cell.border = cell_border

    # Write Excel formula for Dial Diff = Proposed_Dial - Current_Dial
    if diff_col in df.columns and "Proposed_Dial" in df.columns and "Current_Dial" in df.columns:
        diff_idx = df.columns.get_loc(diff_col) + 1
        proposed_idx = df.columns.get_loc("Proposed_Dial") + 1
        current_idx = df.columns.get_loc("Current_Dial") + 1
        diff_letter = get_column_letter(diff_idx)
        proposed_letter = get_column_letter(proposed_idx)
        current_letter = get_column_letter(current_idx)

        for row_idx in range(data_start_row, ws.max_row + 1):
            ws[f"{diff_letter}{row_idx}"] = f"={proposed_letter}{row_idx}-{current_letter}{row_idx}"

    # Conditional formatting for Dial Diff (New - Current) with fixed range
    if diff_col in df.columns:
        diff_idx = df.columns.get_loc(diff_col) + 1
        col_letter = get_column_letter(diff_idx)
        data_range = f"{col_letter}{data_start_row}:{col_letter}{ws.max_row}"
        color_rule = ColorScaleRule(
            start_type="num",
            start_value=-0.5,
            start_color="F8696B",
            mid_type="num",
            mid_value=0,
            mid_color="FFFFFF",
            end_type="num",
            end_value=0.5,
            end_color="63BE7B",
        )
        ws.conditional_formatting.add(data_range, color_rule)

    # Conditional formatting for Loan Num (low -> high)
    if "Loan Num" in df.columns:
        loan_idx = df.columns.get_loc("Loan Num") + 1
        col_letter = get_column_letter(loan_idx)
        data_range = f"{col_letter}{data_start_row}:{col_letter}{ws.max_row}"
        loan_rule = ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFFFFF",
            end_type="max",
            end_color="63BE7B",
        )
        ws.conditional_formatting.add(data_range, loan_rule)

    # Conditional formatting for Ratio columns (0.6 -> 1.4)
    ratio_min = 0.6
    ratio_max = 1.4
    for col_name in df.columns:
        if "Ratio" in col_name:
            ratio_idx = df.columns.get_loc(col_name) + 1
            col_letter = get_column_letter(ratio_idx)
            data_range = f"{col_letter}{data_start_row}:{col_letter}{ws.max_row}"

            # Use midpoint 1.0 for Dialed Model Ratio, otherwise 0.0
            ratio_mid = 1.0 if col_name == "Dialed Model Ratio" else 0.0

            ratio_rule = ColorScaleRule(
                start_type="num",
                start_value=ratio_min,
                start_color="F8696B",
                mid_type="num",
                mid_value=ratio_mid,
                mid_color="FFFFFF",
                end_type="num",
                end_value=ratio_max,
                end_color="63BE7B",
            )
            ws.conditional_formatting.add(data_range, ratio_rule)


def _build_dial_ratio(
    df_summary: pd.DataFrame,
    error_window: str | None,
) -> pd.DataFrame:
    if df_summary.empty:
        return df_summary

    error_window = normalize_error_window(error_window)
    abs_col = f"{error_window} Error Abs" if error_window else None
    ratio_col = f"{error_window} Error Ratio" if error_window else None

    if abs_col and ratio_col and abs_col in df_summary.columns and ratio_col in df_summary.columns:
        ratio_vals = df_summary[ratio_col]
        denom = (ratio_vals - 1).where((ratio_vals - 1) != 0, np.nan)
        df_summary["Dial"] = 1 / ratio_vals
        df_summary["Actual"] = df_summary[abs_col] / denom
        df_summary["Model"] = df_summary["Actual"] * ratio_vals
    else:
        raise ValueError(
            f"Missing '{abs_col}' or '{ratio_col}' for Dial/Actual/Model."
        )

    key_cols = [c for c in ["Sheet", "Bucket_Type", "Status", "Transition", "Bucket"] if c in df_summary.columns]
    ref_cols = [c for c in ["Avg Bal", "Loan Num"] if c in df_summary.columns] # , "MV(MM)", "WALA", "WAC", "FICO", "OCLTV"

    df_norm = df_summary.copy()
    df_norm["Report_Norm"] = df_norm["Report"].astype(str).str.strip().str.lower()

    sentinel = "__NA__"
    for col in key_cols:
        df_norm[col] = df_norm[col].fillna(sentinel)

    ratio_col = f"{error_window} Error Ratio" if error_window else None
    extra_cols = []
    if ratio_col and ratio_col in df_norm.columns:
        extra_cols.append(ratio_col)

    dialed_cols = ["Model"] + (["Actual"] if "Actual" in df_norm.columns else []) + ref_cols + extra_cols
    rename_map = {"Model": "Model_Dialed", "Actual": "Actual"}
    if ratio_col and ratio_col in df_norm.columns:
        rename_map[ratio_col] = "Dialed Model Ratio"
    dialed_df = df_norm[df_norm["Report_Norm"] == "dialed"][key_cols + dialed_cols].rename(
        columns=rename_map
    )
    undialed_df = df_norm[df_norm["Report_Norm"] == "undialed"][key_cols + ["Model"]].rename(
        columns={"Model": "Model_Undialed"}
    )

    if dialed_df.empty or undialed_df.empty:
        available_reports = sorted(df_norm["Report_Norm"].dropna().unique().tolist())
        raise ValueError(
            f"Missing Dialed/Undialed reports for derived dial. Available: {available_reports}"
        )

    df_dial_ratio = dialed_df.merge(undialed_df, on=key_cols, how="inner")
    df_dial_ratio["Current_Dial"] = df_dial_ratio["Model_Dialed"] / df_dial_ratio["Model_Undialed"]
    df_dial_ratio["Implied_Dial"] = df_dial_ratio["Actual"] / df_dial_ratio["Model_Undialed"]
    df_dial_ratio["Proposed_Dial"] = df_dial_ratio["Implied_Dial"]
    diff_col = "Dial Diff (New - Current)"
    df_dial_ratio[diff_col] = df_dial_ratio["Proposed_Dial"] - df_dial_ratio["Current_Dial"]

    for col in key_cols:
        df_dial_ratio[col] = df_dial_ratio[col].replace(sentinel, np.nan)

    output_cols = key_cols + ref_cols + [
        "Model_Dialed",
        "Model_Undialed",
        "Actual",
        "Dialed Model Ratio",
        "Current_Dial",
        "Implied_Dial",
        "Proposed_Dial",
        diff_col,
    ]
    output_cols = [c for c in output_cols if c in df_dial_ratio.columns]
    return df_dial_ratio[output_cols]


def _build_summary(dealtype: str) -> pd.DataFrame:
    summary_rows: list[dict] = []

    for report_label, base_dir in BASE_DIRS.items():
        if MANUAL_EXCEL_PATHS.get(report_label):
            excel_path = Path(MANUAL_EXCEL_PATHS[report_label])
        else:
            excel_path = find_latest_tracking_file(dealtype, base_dir)

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        print(f"[{dealtype}] [{report_label}] Using file: {excel_path}")

        rows = extract_summary_rows(
            excel_path=excel_path,
            bucket_type=BUCKET_TYPE,
            status_sheets=STATUS_SHEETS,
            exclude_sheets=EXCLUDE_SHEETS,
            verbose=False,
        )

        for row in rows:
            row["Report"] = report_label
        summary_rows.extend(rows)

    return pd.DataFrame(summary_rows)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    results: dict[str, pd.DataFrame] = {}
    window_label = normalize_error_window(ERROR_WINDOW) or "ALL"

    for dealtype in DEALTYPES:
        try:
            df_summary = _build_summary(dealtype)
            if df_summary.empty:
                print(f"[{dealtype}] No rows found. Skipping.")
                continue

            df_dial_ratio = _build_dial_ratio(df_summary, ERROR_WINDOW)
            results[dealtype] = df_dial_ratio

            print(f"[{dealtype}] Processed.")
        except Exception as exc:
            print(f"[{dealtype}] Failed: {exc}")

    if not results:
        print("No outputs generated.")
        return

    output_path = OUTPUT_DIR / f"dial_ratio_by_deal_{window_label}.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for dealtype, df in results.items():
            df.to_excel(writer, index=False, sheet_name=dealtype)

    wb = load_workbook(output_path)
    for dealtype, df in results.items():
        ws = wb[dealtype]
        _apply_excel_formatting(ws, df, "Dial Diff (New - Current)", sheet_title=dealtype)
    wb.save(output_path)

    print(f"Saved multi-sheet output: {output_path}")


if __name__ == "__main__":
    main()
